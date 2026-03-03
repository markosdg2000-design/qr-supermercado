#!/usr/bin/env python3
"""Importa Excel/CSV y genera el JSON DATA para grandes o medianos.

Uso rápido:
  python scripts/import_excel.py grandes --input datos_grandes.xlsx --output data_grandes.json
  python scripts/import_excel.py medianos --input datos_medianos.xlsx --output data_medianos.json --station ES05

Opcionalmente, puedes actualizar directamente el HTML:
  python scripts/import_excel.py grandes --input datos.xlsx --update-html grandes.html
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Iterable


def normalize_header(text: str) -> str:
    s = unicodedata.normalize("NFKD", str(text or ""))
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.upper().replace("\n", " ").replace("\r", " ").strip()
    s = re.sub(r"\s+", "", s)
    s = s.replace("º", "N").replace("°", "N").replace("N.", "N")
    return s


def clean_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def read_csv_rows(path: Path) -> list[dict[str, Any]]:
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    sample = raw[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.DictReader(raw.splitlines(), dialect=dialect)
    return [dict(row) for row in reader]


def read_xlsx_rows(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise SystemExit(
            "Para leer .xlsx necesitas openpyxl instalado. "
            "Instala con: pip install openpyxl"
        ) from exc

    wb = load_workbook(path, data_only=True, read_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"La hoja '{sheet_name}' no existe en {path}. Hojas: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    rows = ws.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    keys = [str(h or "") for h in headers]
    out: list[dict[str, Any]] = []
    for row in rows:
        record = {keys[i]: row[i] if i < len(row) else "" for i in range(len(keys))}
        out.append(record)
    return out


def read_rows(path: Path, sheet_name: str | None) -> list[dict[str, Any]]:
    ext = path.suffix.lower()
    if ext in {".csv", ".txt"}:
        return read_csv_rows(path)
    if ext in {".xlsx", ".xlsm"}:
        return read_xlsx_rows(path, sheet_name)
    raise SystemExit(f"Formato no soportado: {path}. Usa CSV o XLSX.")


def get_first(record: dict[str, Any], aliases: Iterable[str]) -> str:
    normalized = {normalize_header(k): v for k, v in record.items()}
    for alias in aliases:
        v = normalized.get(normalize_header(alias))
        if clean_value(v):
            return clean_value(v)
    return ""


def normalize_station(raw_station: str, fallback: str) -> str:
    s = clean_value(raw_station).upper()
    if not s:
        return fallback
    m = re.search(r"ES\s*0?([0-9]{1,2})", s)
    if m:
        return f"ES{int(m.group(1)):02d}"
    m2 = re.search(r"0?([0-9]{1,2})", s)
    if m2:
        return f"ES{int(m2.group(1)):02d}"
    return fallback


def build_grandes(rows: list[dict[str, Any]]) -> dict[str, Any]:
    models: dict[str, dict[str, list[dict[str, str]]]] = defaultdict(lambda: defaultdict(list))

    for r in rows:
        model = get_first(r, ["COCHE", "MODELO", "MODEL"])
        qr = get_first(r, ["COMPONENTE", "QR", "CODIGOQR", "CODIGO", "COMPONENT"])
        if not model or not qr:
            continue

        entry = {
            "Operacion": get_first(r, ["TEXTOBREVEDEMATERIAL", "OPERACION", "MATERIAL", "TEXTOBREVE", "DESCRIPCION"]),
            "Ubicacion": get_first(r, ["UBICACIONDESEADA", "UBICACION", "UBICACIONDESEADA"]),
            "BloqueBol": get_first(r, ["BLOQUEBOL", "BLOQUE", "BOL"]),
            "InstruccionPrep": get_first(r, ["DESCRIPCIONDECOMOHAYQUEPREPARARELBLOQUE", "INSTRUCCIONPREP", "PREPARACION", "DESCRIPCIONPREP"]),
        }

        dia_sec = get_first(r, ["DIASEC", "DIASECUENCIA", "DIA SEC.", "DIA SEC", "DIA_SECUENCIA"])
        if dia_sec:
            entry["DiaSec"] = dia_sec

        if entry not in models[model][qr]:
            models[model][qr].append(entry)

    return {
        "generated_utc": datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z"),
        "models": models,
    }


def build_medianos(rows: list[dict[str, Any],], default_station: str) -> dict[str, Any]:
    data: dict[str, dict[str, dict[str, Any]]] = defaultdict(lambda: defaultdict(dict))

    for r in rows:
        station = normalize_station(get_first(r, ["ESTACION", "AGRUP", "AREA"]), default_station)
        model = get_first(r, ["COCHE", "MODELO", "MODEL"])
        if not model:
            continue

        component = get_first(r, ["MATERIAL", "COMPONENTE_PADRE", "COMPONENTEBASE", "COMPONENTE"])
        qr = get_first(r, ["NºSERIE", "N SERIE", "NSERIE", "QR", "CODIGOQR", "COMPONENTE"])
        if not qr:
            continue

        entry = {
            "Operacion": get_first(r, ["TEXTOBREVEDEMATERIAL", "OPERACION", "MATERIAL", "TEXTOBREVE"]),
            "Ubicacion": get_first(r, ["UBICACIONDESEADA", "UBICACION", "ESTACION"]),
            "Maleta": get_first(r, ["MALETA", "BLOQUEBOL", "BLOQUE"]),
        }
        dia_sec = get_first(r, ["DIASEC", "DIASECUENCIA", "DIA SEC.", "DIA SEC", "DIA_SECUENCIA"])
        if dia_sec:
            entry["DiaSec"] = dia_sec

        model_node = data[station][model]

        if component and component != qr:
            comp_node = model_node.setdefault(component, {})
            bucket = comp_node.setdefault(qr, [])
        else:
            bucket = model_node.setdefault(qr, [])

        if entry not in bucket:
            bucket.append(entry)

    return data


def json_default(obj: Any) -> Any:
    if isinstance(obj, defaultdict):
        return dict(obj)
    raise TypeError


def replace_data_in_html(html_path: Path, payload: dict[str, Any]) -> None:
    text = html_path.read_text(encoding="utf-8")
    replacement = "const DATA = " + json.dumps(payload, ensure_ascii=False) + ";"
    new_text, count = re.subn(r"const DATA = .*?;\n", replacement + "\n", text, count=1, flags=re.S)
    if count != 1:
        raise SystemExit(f"No se pudo localizar 'const DATA = ...;' en {html_path}")
    html_path.write_text(new_text, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Genera DATA JSON para la app desde Excel/CSV")
    parser.add_argument("mode", choices=["grandes", "medianos"], help="Tipo de app a generar")
    parser.add_argument("--input", required=True, type=Path, help="Ruta al .xlsx/.csv")
    parser.add_argument("--sheet", default=None, help="Nombre de hoja (solo Excel)")
    parser.add_argument("--output", type=Path, default=None, help="Ruta de salida JSON")
    parser.add_argument("--update-html", type=Path, default=None, help="Actualiza const DATA dentro del HTML")
    parser.add_argument("--station", default="ES05", help="Estación por defecto para medianos (ej. ES05)")
    parser.add_argument("--pretty", action="store_true", help="JSON con indentación")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    rows = read_rows(args.input, args.sheet)

    if args.mode == "grandes":
        payload = build_grandes(rows)
    else:
        payload = build_medianos(rows, default_station=args.station.upper())

    if args.output:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        args.output.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2 if args.pretty else None, default=json_default),
            encoding="utf-8",
        )
        print(f"JSON generado en: {args.output}")

    if args.update_html:
        replace_data_in_html(args.update_html, payload)
        print(f"HTML actualizado: {args.update_html}")

    if not args.output and not args.update_html:
        print(json.dumps(payload, ensure_ascii=False, indent=2 if args.pretty else None, default=json_default))


if __name__ == "__main__":
    main()
